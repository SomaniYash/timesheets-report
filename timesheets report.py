import streamlit as st
import pandas as pd
import io
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

st.set_page_config(page_title="Monthly Hours Report", page_icon="📊", layout="wide")

st.markdown("""
<style>
    .main-title { font-size:2rem; font-weight:700; color:#01696f; margin-bottom:0.25rem; }
    .sub-title  { color:#7a7974; margin-bottom:2rem; }
    .metric-box { background:#f9f8f5; border:1px solid #dcd9d5; border-radius:8px;
                  padding:1rem; text-align:center; }
    .metric-num { font-size:1.8rem; font-weight:700; color:#01696f; }
    .metric-lbl { font-size:0.85rem; color:#7a7974; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">📊 Monthly Hours Report Generator</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Upload weekly CSV files for all employees — get a full monthly Excel report.</div>', unsafe_allow_html=True)

HOUR_COLS = [
    "Regular Hours", "Night Awake Hours", "Night Sleep Hours",
    "On Call Regular", "On Call Night Sleep",
    "Honorarium Hours",
    "Team Lead Regular", "Team Lead Night Awake", "Team Lead Night Sleep", "Team Lead Training",
    "Training Hours",
    "Sick Hours (Regular)", "Sick Hours (Night Awake)", "Sick Hours (Night Sleep)",
    "Vacation Hours (Regular)", "Vacation Hours (Night Awake)", "Vacation Hours (Night Sleep)",
    "Regular Premium Hours", "Night Awake Premium Hours", "Night Sleep Premium Hours",
    "Team Lead Premium", "Team Lead Night Sleep Premium", "Team Lead Night Awake Premium",
]


def safe_float(val):
    try:
        v = float(str(val).strip())
        return 0.0 if (v != v) else v
    except:
        return 0.0


def parse_week_date(s):
    s = re.sub(r'\d{2}:\d{2}:\d{2}\s*GMT.*', '', s).strip()
    for fmt in ["%a %b %d %Y", "%a %b  %d %Y"]:
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except:
            pass
    return None


def normalize_member(name):
    """Normalize member name: collapse spaces, strip On Call prefix."""
    name = re.sub(r'\s+', ' ', name).strip()
    # Remove "On Call " prefix so "On Call Scott Marshall" → "Scott Marshall"
    name = re.sub(r'(?i)^on call\s+', '', name).strip()
    return name


def parse_csv(file_bytes):
    text = file_bytes.decode("utf-8", errors="replace")
    rows = [r.split(",") for r in text.splitlines()]

    employee = ""
    col_map = {}
    data_start_idx = 0

    for i, row in enumerate(rows):
        if not row:
            continue
        cell0 = row[0]

        if "Employee:" in cell0:
            m = re.search(r'Employee:\s*(.+?)(?:undefined|$)', cell0)
            if m:
                employee = m.group(1).strip()

        if re.match(r'^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+\w+\s+\d+\s+\d{4}', cell0.strip()):
            data_start_idx = i
            break

    if rows:
        for j, h in enumerate(rows[0]):
            h2 = h.strip()
            if h2 and h2.lower() != "undefined":
                col_map[h2] = j

    records = []
    current_day = None

    # Determine Start Time column index (used to detect shift rows)
    st_idx = col_map.get("Start Time", 1)
    member_idx = col_map.get("Member Supported", len(rows[0]) - 1 if rows else 0)

    for row in rows[data_start_idx:]:
        if not row:
            continue
        cell0 = row[0].strip()

        if "Total hours" in cell0:
            continue

        # ── New day header: "Sat May 23 2026" ───────────────────────────────
        # IMPORTANT: The first shift of the day is on the SAME ROW as the date header.
        # So after extracting the date, we do NOT skip — we fall through to process
        # the shift data on that same row.
        day_m = re.match(r'^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(\w+)\s+(\d+)\s+(\d{4})', cell0)
        if day_m:
            try:
                current_day = datetime.strptime(
                    f"{day_m.group(2)} {day_m.group(3)} {day_m.group(4)}", "%b %d %Y"
                ).date()
            except:
                pass
            # Fall through — do NOT continue — the row may also contain shift data

        if current_day is None:
            continue

        # ── Detect shift row by Start Time column ────────────────────────────
        # Handles both: day-header rows (first shift inline) and
        # continuation rows (empty cell0, subsequent shifts same day)
        start_val = row[st_idx].strip() if st_idx < len(row) else ""
        if not re.match(r'\d+:\d+', start_val):
            continue

        # ── Member name ───────────────────────────────────────────────────────
        raw_member = row[member_idx].strip() if member_idx < len(row) else ""
        if not raw_member or raw_member.lower() in ("undefined", ""):
            continue
        member = normalize_member(raw_member)

        rec = {"employee": employee, "day": current_day, "member": member}
        for hcol in HOUR_COLS:
            idx = col_map.get(hcol)
            rec[hcol] = safe_float(row[idx]) if (idx is not None and idx < len(row)) else 0.0

        records.append(rec)

    return records, employee


def determine_target_month(records):
    from collections import Counter
    counter = Counter()
    for r in records:
        if r["day"]:
            counter[(r["day"].year, r["day"].month)] += 1
    return counter.most_common(1)[0][0] if counter else (None, None)


def aggregate(all_records):
    if not all_records:
        return None, None, None
    year, month = determine_target_month(all_records)
    if not year:
        return None, None, None
    filtered = [r for r in all_records
                if r["day"] and r["day"].year == year and r["day"].month == month]
    agg = defaultdict(lambda: defaultdict(float))
    for r in filtered:
        key = (r["member"], r["employee"])
        for hcol in HOUR_COLS:
            agg[key][hcol] += r.get(hcol, 0.0)
    return agg, year, month


def build_excel(agg, year, month):
    month_label = datetime(year, month, 1).strftime("%B %Y")
    wb = Workbook()

    TEAL, TEAL_DARK, WHITE = "01696F", "0C4E54", "FFFFFF"
    LIGHT_ROW, TOTAL_FILL = "F3F0EC", "E6E4DF"

    def mk_font(bold=False, color="28251D", size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def mk_fill(color):
        return PatternFill("solid", fgColor=color)

    hdr_font    = mk_font(bold=True,  color=WHITE, size=11)
    member_font = mk_font(bold=True,  color=WHITE, size=12)
    staff_font  = mk_font(bold=True,  size=11)
    data_font   = mk_font(size=10)
    total_font  = mk_font(bold=True,  size=10)
    title_font  = mk_font(bold=True,  color=TEAL, size=15)
    sub_font    = mk_font(color="7a7974", size=10)
    grand_font  = mk_font(bold=True,  color=WHITE, size=11)

    hdr_fill    = mk_fill(TEAL)
    member_fill = mk_fill(TEAL_DARK)
    total_fill  = mk_fill(TOTAL_FILL)
    grand_fill  = mk_fill(TEAL)
    alt_fill    = mk_fill(LIGHT_ROW)

    left   = Alignment(horizontal="left",   vertical="center", indent=1)
    right  = Alignment(horizontal="right",  vertical="center")
    wrap_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_s = Side(style="thin", color="DCD9D5")
    thin_b = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    num_fmt   = '#,##0.0;-#,##0.0;"-"'
    TOTAL_COL = 2 + len(HOUR_COLS) + 1
    LAST_COL  = TOTAL_COL
    FL = get_column_letter(3)
    LL = get_column_letter(2 + len(HOUR_COLS))

    def write_header_row(ws, row_num, first_col_label="Employee"):
        ws.row_dimensions[row_num].height = 42
        for ci, h in enumerate([first_col_label] + HOUR_COLS + ["Total Hours"], start=2):
            c = ws.cell(row=row_num, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = wrap_c; c.border = thin_b
        ws.freeze_panes = f"B{row_num + 1}"

    def set_col_widths(ws):
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions[get_column_letter(2)].width = 30
        for ci in range(3, LAST_COL + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 13

    def write_title(ws, title_text):
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 16
        ws.row_dimensions[4].height = 8
        t = ws.cell(row=2, column=2, value=title_text)
        t.font = title_font; t.alignment = left
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=LAST_COL)
        s = ws.cell(row=3, column=2,
                    value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {month_label}")
        s.font = sub_font; s.alignment = left
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=LAST_COL)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    set_col_widths(ws)
    write_title(ws, f"Monthly Hours Report — {month_label}")

    HDR_ROW = 6
    write_header_row(ws, HDR_ROW, "Employee")

    all_members = sorted(set(m for (m, _) in agg.keys()))
    cur = HDR_ROW + 1

    for member in all_members:
        ws.row_dimensions[cur].height = 22
        mc = ws.cell(row=cur, column=2, value=f"   {member}")
        mc.font = member_font; mc.fill = member_fill; mc.alignment = left
        ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=LAST_COL)
        cur += 1

        employees = sorted(e for (m, e) in agg.keys() if m == member)
        staff_start = cur

        for i, emp in enumerate(employees):
            hours = agg[(member, emp)]
            ws.row_dimensions[cur].height = 18
            rf = alt_fill if i % 2 == 0 else None

            ec = ws.cell(row=cur, column=2, value=emp)
            ec.font = staff_font; ec.alignment = left; ec.border = thin_b
            if rf: ec.fill = rf

            for ci, hcol in enumerate(HOUR_COLS, start=3):
                val = hours.get(hcol, 0.0)
                c = ws.cell(row=cur, column=ci, value=val if val else None)
                c.font = data_font; c.number_format = num_fmt
                c.alignment = right; c.border = thin_b
                if rf: c.fill = rf

            tc = ws.cell(row=cur, column=TOTAL_COL,
                         value=f"=SUM({FL}{cur}:{LL}{cur})")
            tc.font = total_font; tc.number_format = num_fmt
            tc.alignment = right; tc.border = thin_b
            if rf: tc.fill = rf
            cur += 1

        staff_end = cur - 1
        ws.row_dimensions[cur].height = 20
        sl = ws.cell(row=cur, column=2, value=f"Subtotal — {member}")
        sl.font = total_font; sl.fill = total_fill; sl.alignment = left; sl.border = thin_b

        for ci in range(3, LAST_COL + 1):
            cl = get_column_letter(ci)
            c = ws.cell(row=cur, column=ci,
                        value=f"=SUM({cl}{staff_start}:{cl}{staff_end})")
            c.font = total_font; c.number_format = num_fmt
            c.fill = total_fill; c.alignment = right; c.border = thin_b
        cur += 2

    ws.row_dimensions[cur].height = 24
    gl = ws.cell(row=cur, column=2, value="GRAND TOTAL — All Members")
    gl.font = grand_font; gl.fill = grand_fill; gl.alignment = left; gl.border = thin_b
    for ci in range(3, LAST_COL + 1):
        cl = get_column_letter(ci)
        c = ws.cell(row=cur, column=ci,
                    value=f"=SUM({cl}{HDR_ROW + 1}:{cl}{cur - 1})")
        c.font = grand_font; c.number_format = num_fmt
        c.fill = grand_fill; c.alignment = right; c.border = thin_b

    # ── Per-employee sheets ───────────────────────────────────────────────────
    all_employees = sorted(set(e for (_, e) in agg.keys()))

    for emp in all_employees:
        safe = re.sub(r'[\\/*?:\[\]]', "", emp)[:31]
        ws_e = wb.create_sheet(title=safe)
        set_col_widths(ws_e)
        write_title(ws_e, f"{emp} — {month_label}")

        HDR = 6
        write_header_row(ws_e, HDR, "Member Supported")

        emp_pairs = sorted(
            [(m, agg[(m, emp)]) for (m, e) in agg.keys() if e == emp],
            key=lambda x: x[0]
        )
        r = HDR + 1
        for i, (member, hours) in enumerate(emp_pairs):
            ws_e.row_dimensions[r].height = 18
            rf = alt_fill if i % 2 == 0 else None

            mc = ws_e.cell(row=r, column=2, value=member)
            mc.font = staff_font; mc.alignment = left; mc.border = thin_b
            if rf: mc.fill = rf

            for ci, hcol in enumerate(HOUR_COLS, start=3):
                val = hours.get(hcol, 0.0)
                c = ws_e.cell(row=r, column=ci, value=val if val else None)
                c.font = data_font; c.number_format = num_fmt
                c.alignment = right; c.border = thin_b
                if rf: c.fill = rf

            tc = ws_e.cell(row=r, column=TOTAL_COL,
                           value=f"=SUM({FL}{r}:{LL}{r})")
            tc.font = total_font; tc.number_format = num_fmt
            tc.alignment = right; tc.border = thin_b
            if rf: tc.fill = rf
            r += 1

        ws_e.row_dimensions[r].height = 22
        tl = ws_e.cell(row=r, column=2, value=f"Total — {emp}")
        tl.font = grand_font; tl.fill = grand_fill; tl.alignment = left; tl.border = thin_b
        for ci in range(3, LAST_COL + 1):
            cl = get_column_letter(ci)
            c = ws_e.cell(row=r, column=ci,
                          value=f"=SUM({cl}{HDR + 1}:{cl}{r - 1})")
            c.font = grand_font; c.number_format = num_fmt
            c.fill = grand_fill; c.alignment = right; c.border = thin_b

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, month_label


# ══════════════════════════════════════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════════════════════════════════════
uploaded = st.file_uploader(
    "Drop CSV or ZIP files here (weekly exports — multiple employees, multiple weeks at once)",
    type=["csv", "zip"],
    accept_multiple_files=True,
)

def extract_csv_files(uploaded_files):
    """Return list of (filename, bytes) tuples from CSVs and ZIPs."""
    import zipfile
    csv_files = []
    for f in uploaded_files:
        if f.name.lower().endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(f.read())) as zf:
                    for name in zf.namelist():
                        if name.lower().endswith(".csv") and not name.startswith("__MACOSX"):
                            csv_files.append((name, zf.read(name)))
            except Exception as ex:
                st.warning(f"Could not open ZIP `{f.name}`: {ex}")
        else:
            csv_files.append((f.name, f.read()))
    return csv_files

if uploaded:
    raw_csv_files = extract_csv_files(uploaded)

    if not raw_csv_files:
        st.error("No CSV files found in the uploaded files/ZIPs.")
        st.stop()

    st.caption(f"Found **{len(raw_csv_files)}** CSV file(s) across {len(uploaded)} uploaded file(s)")

    all_records = []
    parse_errors = []
    progress = st.progress(0)

    for idx, (fname, fbytes) in enumerate(raw_csv_files):
        try:
            recs, _ = parse_csv(fbytes)
            all_records.extend(recs)
        except Exception as ex:
            parse_errors.append(f"`{fname}`: {ex}")
        progress.progress((idx + 1) / len(raw_csv_files))
    progress.empty()

    for e in parse_errors:
        st.warning(f"Parse issue — {e}")

    if not all_records:
        st.error("No valid shift records found. Check your CSV files.")
        st.stop()

    agg, year, month = aggregate(all_records)
    if not agg:
        st.error("Could not determine target month.")
        st.stop()

    month_label = datetime(year, month, 1).strftime("%B %Y")
    unique_employees = sorted(set(e for (_, e) in agg.keys()))
    unique_members   = sorted(set(m for (m, _) in agg.keys()))
    total_hours = sum(sum(h.get(hc, 0) for hc in HOUR_COLS) for h in agg.values())

    st.markdown(f"### Report Period: **{month_label}**")
    c1, c2, c3, c4 = st.columns(4)
    for col, num, label in [
        (c1, len(uploaded), "CSV Files Uploaded"),
        (c2, len(unique_employees), "Staff Members"),
        (c3, len(unique_members), "Members Supported"),
        (c4, f"{total_hours:,.1f}", "Total Hours"),
    ]:
        with col:
            st.markdown(
                f'<div class="metric-box"><div class="metric-num">{num}</div>'
                f'<div class="metric-lbl">{label}</div></div>',
                unsafe_allow_html=True,
            )

    st.markdown("---")

    with st.expander("Preview Aggregated Data", expanded=False):
        rows = []
        for (member, emp), hours in agg.items():
            row = {"Member Supported": member, "Employee": emp}
            for hcol in HOUR_COLS:
                v = hours.get(hcol, 0)
                row[hcol] = v if v else "-"
            row["Total"] = sum(hours.get(hc, 0) for hc in HOUR_COLS)
            rows.append(row)
        preview_df = pd.DataFrame(rows).sort_values(["Member Supported", "Employee"])
        st.dataframe(preview_df, use_container_width=True, hide_index=True)

    st.markdown("### Download Report")
    if st.button("Generate Excel Report", type="primary", use_container_width=True):
        with st.spinner("Building Excel..."):
            buf, ml = build_excel(agg, year, month)
        fname = f"Monthly_Hours_{ml.replace(' ', '_')}.xlsx"
        st.download_button(
            label=f"Download {fname}",
            data=buf,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.success("Report ready! Click above to download.")

else:
    st.info("Upload one or more CSV files to get started.")
    with st.expander("How to use this app"):
        st.markdown("""
        1. Export **weekly timesheet CSVs** from your system for each employee
        2. Upload **all CSVs at once** — multiple weeks and multiple employees together
        3. The app auto-detects the **target month** (whichever month has the most shift days wins)
        4. Click **Generate Excel Report** and download

        **Excel output:**
        - **Summary sheet** — Each member (Don Mackenzie, Shawn Billy, etc.) as a section,  
          listing every employee who worked with them and hours by type
        - **Per-employee sheets** — One tab per staff member with their breakdown by member
        - Zero values display as **-** for clean reading
        - Subtotals per member + Grand Total row
        """)
